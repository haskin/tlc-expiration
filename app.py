from tkinter import *
from tkinter import ttk
import requests
from openpyxl import load_workbook
import time

class TlcDrivLicExpiration():
    
    def __init__(self):
        '''Initializes all important variables and runs program.
        '''
        ################# Tkinter UI/UX
        # Root
        self.root = Tk()
        self.root.minsize(450,300)
        self.root.bind('<Return>', self.test_drivers)
        self.root.title("TLC Driver's License Status Check")
        self.instr = ttk.Label(self.root)

        # Instructions
        self.instr.configure(text='Please have an excel file named "license.xlsx" in the root directory of the program.'
                    + '\nLicenses must be in a column labeled "TLC licence number" in order to be checked.')
        self.instr.grid()

        # Progress Bar
        self.prog = ttk.Progressbar(self.root, orient=HORIZONTAL, length=450, mode='determinate')
        self.prog.grid(sticky =(W, E))
        #Enter Button
        ttk.Button(self.root, text="Find Expired Licenses", command=self.test_drivers).grid(sticky = E)

        # Expired Driver Licenses
        ttk.Label(self.root, text="Expired Driver's Licenses:").grid(sticky=W)
        self.results = StringVar()
        ttk.Label(self.root, textvariable=self.results).grid(sticky=W)

        for child in self.root.winfo_children(): child.grid_configure(padx=5, pady=5)
        ################# Tkinter UI/UX END

        ################# Excel variables
        self.wb = load_workbook(filename = 'expired.xlsx')
        #Dictionary for column names to column index
        self.cols = {}
        self.set_cols()
        self.expired = []
        self.driv_lic_amount = None
        self.set_driv_lic_amount()
        ################# Excel variables

        
        self.root.mainloop()
    
    def set_driv_lic_amount(self):
        '''Sets the amount of driver licenses.
        '''
        ws = self.wb.active
        # self.driv_lic_amount = sum([1 for w in ws[self.cols['TLC licence number']] != "123"])
        self.driv_lic_amount = len(ws[self.cols['TLC licence number']])
        self.prog["maximum"] = self.driv_lic_amount

    def set_cols(self):
        '''Sets the self.cols DICT to map from column names
        to the letter which the column is in Excel.
        E.g. 1st column name == "A"
        '''
        #Active worksheet from Excel workbook
        ws = self.wb.active
        unic = 65 #Unicode for "A"
        indx = 1
        while True:
            col_name = ws.cell(row=1, column=indx).value
            if not col_name or indx > 100: 
                break
            else:
                self.cols[col_name] = chr(unic)
                unic += 1
                indx += 1
        pass

    def create_URL(self, license):
        '''Returns URL to connect to NYC OpenData API.
        '''
        return "https://data.cityofnewyork.us/api/id/xjfq-wh2d.json?$query=select *%2C %3Aid search '{}' limit 100&$$query_timeout_seconds=60".format(license)

    def find_expired(self, license):
        '''Sends a request to NYC OpenData API to get JSON data
        on the driver's license. If no data is found, appends
        self.expired with the license. 

        Updates the results that appear in Window with the 
        expired licenses.
        '''
        url = self.create_URL(license)
        try: 
            res = requests.get(url)
            res_license = res.json()[0]['license_number']
            if str(license) == res_license:
                return
            else: 
                raise Error 
        except:
            self.expired.append(license)
            self.results.set([i for i in self.expired])

    def test_drivers(self, *args):
        '''Goes through each license found in the 'TLC licence number'
        column and sends it to self.find_expired(). Updates the progress
        bar with each driver.
        '''
        ws = self.wb.active
        for license in ws[self.cols['TLC licence number']]:
            #Avoids header
            if 'TLC licence number' == license.value:
                continue
            if license.value != 123:
                # self.prog["value"] += 1
                time.sleep(0.05)
                self.prog["value"] += 1
                self.prog.update()
                self.find_expired(license.value)
                # self.prog.update()
        #Sets the progress bar to 100%
        time.sleep(0.05)
        self.prog["value"] = self.driv_lic_amount
        self.prog.update()

test = TlcDrivLicExpiration()